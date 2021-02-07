from PyQt5.QtCore import pyqtSignal, QThread
from PyQt5.QtWidgets import QApplication
import sys
import MySQLdb
from openpyxl import Workbook
import pandas as pd
from openpyxl.utils import get_column_letter
import openpyxl


class WorkThread(QThread):
    trigger = pyqtSignal(str)

    def __init__(self):
        super(WorkThread, self).__init__()
        self.host = "127.0.0.1"
        self.port = 3306
        self.user = "root"
        self.passwd = "123456"
        self.db = "mydatabase"
        self.key = str
        self.table_all = str
        self.table = str

    def set(self, db, key, table_all, table):  # 设置传入参数函数
        # 数据库名称指定
        self.db = db
        # 列表行所数据表
        self.table_all = table_all
        # 查询数据所在数据表
        self.table = table
        # 需要查询列的名称
        self.key = key

    def run(self):
        sql = 'SELECT * FROM {}'.format(self.table_all)
        self.read_mysql_to_xlsx(sql, 'all.xlsx')
        df = pd.read_excel('all.xlsx')
        for excelname in df[self.key]:
            print(excelname)
            sql = 'SELECT * FROM {} where {} = \'{}\''.format(self.table, self.key, excelname)
            self.read_mysql_to_xlsx(sql, str(excelname))
        self.trigger.emit('完成')  # 传回线程数据

    def query_all(self, sql):
        # 连接数据库，查询数据
        db = MySQLdb.connect(host=self.host, port=self.port, user=self.user, passwd=self.passwd, db=self.db, charset="GBK")
        # 使用cursor()方法获取操作游标
        cur = db.cursor()
        # 使用execute方法执行SQL语句
        cur.execute(sql)  # 返回受影响的行数
        fields = [field[0] for field in cur.description]  # 获取所有字段名
        all_data = cur.fetchall()  # 所有数据
        print('数据行数：%d' % len(all_data))
        print('##'*20)
        return all_data, fields

    def read_mysql_to_xlsx(self, sql, excelname):
        # 循环数据写入内容
        jb_date_lists = self.query_all(sql)
        # print(jb_date_lists)
        jb_date_list = jb_date_lists[0]
        descripte = jb_date_lists[1]
        # 要创建的xlsx名称
        # dest_filename = excelname + '.xlsx'
        dest_filename = excelname
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "data"
        # 列名
        for i in range(0, len(descripte)):
            ws1.cell(row=1, column=i + 1, value=descripte[i])
        # 写入数据
        for i in range(2, len(jb_date_list) + 2):
            for j in range(0, len(descripte)):
                if jb_date_list[i - 2][j] is None:
                    ws1.cell(row=i, column=j + 1, value='')
                else:
                    ws1.cell(row=i, column=j + 1, value=jb_date_list[i - 2][j])
        # 创建xlsx
        wb.save(filename=dest_filename)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    t = WorkThread()
    t.set('mydatabase', '来源', 'download_times', 'comparison')
    t.run()
